# -*- coding:utf-8 -*-
# edit by fuzongfei

import datetime
import difflib
import hashlib
import json
import os
import pyminizip
import random
import string
import time
from datetime import datetime

import mysql.connector as mdb
import pymysql
from asgiref.sync import async_to_sync
from celery import shared_task
from channels.layers import get_channel_layer
from django.core.cache import cache
from django.core.files import File
from django.core.mail import EmailMessage
from django.db.models import F
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from AuditSQL.settings import EMAIL_FROM
from ProjectManager.inception.inception_api import IncepSqlCheck
from ProjectManager.models import OnlineAuditContents, OnlineAuditContentsReply, MonitorSchema, IncepMakeExecTask, \
    DomainName, InceptionHostConfig, DataExport, Files
from ProjectManager.utils import update_tasks_status
from UserManager.models import UserAccount
from UserManager.utils import GetEmailAddr

channel_layer = get_channel_layer()


@shared_task
def send_commit_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(OnlineAuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operate_dba')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向_commit_mail.html渲染data数据
    if DomainName.objects.filter().first():
        domain_name = DomainName.objects.get().domain_name
    record = OnlineAuditContents.objects.annotate(group_name=F('group__group_name')).get(pk=latest_id)
    email_html_body = render_to_string('_send_commit_mail.html', {'data': record, 'domain_name': domain_name})

    # 发送邮件
    msg = EmailMessage(subject=record.title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       )
    msg.content_subtype = "html"

    msg.send()


@shared_task
def send_verify_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(OnlineAuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operate_dba')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向mail_template.html渲染data数据
    record = OnlineAuditContents.objects.get(pk=latest_id)
    email_html_body = render_to_string('_send_verify_mail.html', {
        'data': record,
        'user_role': kwargs['user_role'],
        'username': kwargs['username'],
        'addition_info': kwargs['addition_info']
    })

    # 发送邮件
    headers = {'Reply: ': receiver}
    title = 'Re: ' + record.title
    msg = EmailMessage(subject=title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       headers=headers)
    msg.content_subtype = "html"
    msg.send()


@shared_task
def send_reply_mail(**kwargs):
    latest_id = kwargs['latest_id']
    reply_id = kwargs['reply_id']
    userinfo = GetEmailAddr(OnlineAuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operate_dba')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    title = OnlineAuditContents.objects.get(pk=reply_id).title
    reply_record = OnlineAuditContentsReply.objects.get(pk=latest_id)

    # 向mail_template.html渲染data数据
    email_html_body = render_to_string('_send_reply_mail.html', {
        'reply_record': reply_record,
        'username': kwargs['username'],
    })

    # 发送邮件
    headers = {'Reply: ': receiver}
    title = 'Re: ' + title
    msg = EmailMessage(subject=title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       headers=headers)
    msg.content_subtype = "html"
    msg.send()


def connect_db(**kwargs):
    config = {
        'user': kwargs['user'],
        'password': kwargs['password'],
        'host': kwargs['host'],
        'port': kwargs['port'],
        'database': kwargs['database'],
        'raw': 'True'
    }

    return mdb.connect(**config)


@shared_task
def schema_modify_monitor(**kwargs):
    check_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = connect_db(**kwargs)
    cursor = conn.cursor(dictionary=True)

    query_info = "select table_schema,table_name,group_concat(COLUMN_NAME) as column_name," \
                 "group_concat(COLUMN_DEFAULT) as column_default,group_concat(IS_NULLABLE) as is_nullable," \
                 "group_concat(DATA_TYPE) as data_type,group_concat(CHARACTER_MAXIMUM_LENGTH) as char_length," \
                 "group_concat(COLUMN_TYPE) as column_type,group_concat(COLUMN_COMMENT) as column_comment " \
                 "from columns where table_schema='{schema}' " \
                 "group by table_schema,table_name".format(schema=kwargs['schema'])

    cursor.execute(query_info)

    source_info = []
    table_list = []
    diff_old_data = ''
    diff_new_data = ''
    table_change_data = []

    for row in cursor.fetchall():
        table_schema = row['table_schema']
        table_name = row['table_name']

        md5_source = ''.join(str(row.values()))
        md5_sum = hashlib.md5(md5_source.encode('utf8')).hexdigest()
        source_info.append({'table_schema': table_schema, 'table_name': table_name, 'md5_sum': md5_sum})
        table_list.append(table_name)

    # 如果当前库没有记录，则进行初始化全量同步
    if MonitorSchema.objects.filter(table_schema=kwargs['schema']).first() is None:
        for row in source_info:
            table_schema = row['table_schema']
            table_name = row['table_name']

            query_table_stru = "show create table {}".format('.'.join((table_schema, table_name)))
            cursor.execute(query_table_stru)
            for i in cursor:
                table_stru = i['Create Table']
                row['table_stru'] = str(table_stru)
                MonitorSchema.objects.create(**row)
    else:
        # 如果存在，开始核验数据
        old_data = list(
            MonitorSchema.objects.filter(table_schema=kwargs['schema']).values_list('table_name', flat=True))
        new_data = table_list

        # 找出已删除的表，并处理
        table_remove = list(set(old_data).difference(set(new_data)))
        if table_remove:
            table_change_data.append({'remove': table_remove})
            # 从本地库中删除该表的记录
            MonitorSchema.objects.filter(table_schema=kwargs['schema']).filter(table_name__in=table_remove).delete()

        # 找出新增的表，并处理
        table_add = list(set(new_data).difference(set(old_data)))
        if table_add:
            for i in table_add:
                for j in source_info:
                    if i in j.values():
                        table_change_data.append({'add': j})
                        table_schema = j['table_schema']
                        table_name = j['table_name']
                        query_table_stru = "show create table {}".format('.'.join((table_schema, table_name)))
                        cursor.execute(query_table_stru)
                        for x in cursor:
                            table_stru = x['Create Table']
                            j['table_stru'] = str(table_stru)
                            MonitorSchema.objects.create(**j)

        # 找出相同的表，并核验表结构
        table_intersection = list(set(old_data).intersection(set(new_data)))
        for row in source_info:
            table_schema = row['table_schema']
            table_name = row['table_name']
            new_md5_sum = row['md5_sum']

            if table_name in table_intersection:
                old_table = MonitorSchema.objects.get(table_schema=table_schema, table_name=table_name)
                if new_md5_sum != old_table.md5_sum:
                    query_table_stru = "show create table {}".format('.'.join((table_schema, table_name)))
                    cursor.execute(query_table_stru)
                    for i in cursor:
                        table_stru = i['Create Table']
                        diff_old_data += old_table.table_stru + '\n' * 3
                        diff_new_data += table_stru + '\n' * 3
                        # 更新新表表结构到本地
                        MonitorSchema.objects.update_or_create(table_schema=table_schema, table_name=table_name,
                                                               defaults={'table_stru': table_stru,
                                                                         'md5_sum': new_md5_sum})

    if (diff_old_data and diff_new_data) or table_change_data:
        html_data = ''
        if diff_old_data and diff_new_data:
            diff_data = difflib.HtmlDiff(tabsize=2)
            old_table_stru = list(diff_old_data.split('\n'))
            new_table_stru = list(diff_new_data.split('\n'))
            html_data = diff_data.make_file(old_table_stru, new_table_stru, '旧表-表结构', '新表-表结构', context=False,
                                            numlines=5)

        email_html_body = render_to_string('_monitor_table.html',
                                           {'html_data': html_data, 'table_change_data': table_change_data})
        title = '{db}库表变更[来自:{host},检测时间:{check_time}]'.format(db=kwargs['schema'], host=kwargs['describle'],
                                                               check_time=check_time)
        msg = EmailMessage(subject=title,
                           body=email_html_body,
                           from_email=EMAIL_FROM,
                           to=kwargs['receiver'].split(','),
                           )
        msg.content_subtype = "html"
        msg.send()
    cursor.close()
    conn.close()


"""
status = 0: 推送执行结果
statu = 1: 推送执行进度
"""


@shared_task
def get_osc_percent(user, id, redis_key=None, sqlsha1=None):
    obj = IncepMakeExecTask.objects.get(id=id)
    if sqlsha1 is None:
        sqlsha1 = obj.sqlsha1

    host = obj.dst_host
    database = obj.dst_database

    while True:
        if redis_key in cache:
            data = cache.get(redis_key)
            if data == 'start':
                sql = f"inception get osc_percent '{sqlsha1}'"
                incep_of_audit = IncepSqlCheck(sql, host, database, user)

                # 执行SQL
                incep_of_audit.run_status(1)

                # 每2s获取一次
                time.sleep(2)

            elif data == 'end':
                # 删除key
                cache.delete_pattern(redis_key)
                break
        else:
            break


@shared_task
def incep_async_tasks(user, redis_key=None, sql=None, id=None, exec_status=None):
    obj = IncepMakeExecTask.objects.get(id=id)
    if sql is None:
        sql = obj.sql_content + ';'

    host = obj.dst_host
    database = obj.dst_database

    incep_of_audit = IncepSqlCheck(sql, host, database, user)

    # 执行SQL
    exec_result = incep_of_audit.run_exec(0)

    # 告诉获取进度的线程退出
    cache.set(redis_key, 'end')

    # 更新任务进度
    update_tasks_status(id=id, exec_result=exec_result, exec_status=exec_status)


@shared_task
def stop_incep_osc(user, redis_key=None, id=None):
    obj = IncepMakeExecTask.objects.get(id=id)

    exec_status = None
    if obj.exec_status == '2':
        sqlsha1 = obj.sqlsha1
        exec_status = 0
    elif obj.exec_status == '3':
        sqlsha1 = obj.rollback_sqlsha1
        exec_status = 1

    host = obj.dst_host
    database = obj.dst_database
    sql = f"inception stop alter '{sqlsha1}'"

    # 执行SQL
    incep_of_audit = IncepSqlCheck(sql, host, database, user)
    incep_of_audit.run_status(0)

    # 告诉获取进度的线程退出
    cache.set(redis_key, 'end')

    # 更新任务进度
    update_tasks_status(id=id, exec_status=exec_status)


@shared_task
def make_export_file(user, id):
    obj = DataExport.objects.get(pk=id)
    config = InceptionHostConfig.objects.get(host=obj.dst_host, is_enable=0)

    conn = pymysql.connect(host=config.host,
                           user=config.user,
                           password=config.password,
                           port=config.port,
                           database=obj.dst_database,
                           charset="utf8")

    # 创建目录和生成文件名
    tmp_file = f'media/tmp/{obj.title}.{obj.file_format}'
    zip_file = tmp_file + '.zip'

    try:
        sql = obj.sql_contents
        # 获取内容
        with conn.cursor() as cursor:
            cursor.execute(sql)

            c_result = []

            for row in cursor.fetchall():
                c_result.append(row)

        # 获取标题
        conn.cursorclass = pymysql.cursors.DictCursor
        with conn.cursor() as cursor:
            cursor.execute(sql)

            c_title = []

            for key in cursor.fetchone():
                c_title.append(key)

        wb = Workbook()
        wb.encoding = f'{obj.file_coding}'
        ws = wb.active
        font = Font(name='Courier', size=14)
        align = Alignment(horizontal='right', vertical='center')

        # 写入数据
        ws.append(c_title)
        for irow in c_result:
            ws.append(irow)

        # 设置表格的样式
        for row in range(1, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                ws.cell(row=row, column=column).font = font
                ws.cell(row=row, column=column).alignment = align
                ws.row_dimensions[row].height = 18
                column_alias = ws.cell(row=row, column=column).column
                ws.column_dimensions[f'{column_alias}'].width = 15

        wb.save(tmp_file)

        # 压缩并加密，随机生成12位长度的字符串
        salt = ''.join(random.sample(string.ascii_letters + string.digits, 12))
        pyminizip.compress_multiple([tmp_file], zip_file, salt, 4)

        # 将文件转换为File对象，便于存储
        with open(zip_file, 'rb') as f:
            myfile = File(f)

            # 写入信息到表中
            Files.objects.create(
                export_id=id,
                file_name=obj.title + '.xlsx.gz',
                file_size=os.path.getsize(zip_file),
                files=myfile,
                encryption_key=salt,
                content_type=obj.file_format
            )

        # 删除文件，保留加密和压缩后的文件
        os.remove(tmp_file)
        os.remove(zip_file)

        # 更新进度
        DataExport.objects.filter(pk=id).update(status='2')

        send_data_export_reply_mail.delay(latest_id=id)
    except conn.InternalError as err:
        pull_msg = {'status': 400, 'msg': str(err)}
        DataExport.objects.filter(pk=id).update(status='0')
        async_to_sync(channel_layer.group_send)(user, {"type": "user.message",
                                                       'text': json.dumps(pull_msg)})

    finally:
        conn.close()


@shared_task
def send_data_export_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(DataExport, latest_id)

    receiver = userinfo.get_user_email('proposer', 'operate_dba')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向_send_data_export_mail.html渲染data数据
    if DomainName.objects.filter().first():
        domain_name = DomainName.objects.get().domain_name
    record = DataExport.objects.annotate(group_name=F('group__group_name')).get(pk=latest_id)
    email_html_body = render_to_string('_send_data_export_mail.html', {'data': record, 'domain_name': domain_name})

    # 发送邮件
    msg = EmailMessage(subject=record.title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       )
    msg.content_subtype = "html"
    msg.send()


@shared_task
def send_data_export_reply_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(DataExport, latest_id)

    obj = DataExport.objects.get(pk=latest_id)
    user_list = [obj.proposer, obj.operate_dba]
    receiver = list(UserAccount.objects.filter(username__in=user_list).values_list('email', flat=True))
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向mail_template.html渲染data数据
    if DomainName.objects.filter().first():
        domain_name = DomainName.objects.get().domain_name
    record = Files.objects.get(export_id=latest_id)
    title = DataExport.objects.get(pk=latest_id).title
    email_html_body = render_to_string('_send_data_export_reply_mail.html', {
        'data': record,
        'domain_name': domain_name,
        'file_size_mb': int(record.file_size / 1024 / 1024),
    })

    # 发送邮件
    headers = {'Reply: ': receiver}
    title = 'Re: ' + title
    msg = EmailMessage(subject=title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       headers=headers)
    msg.content_subtype = "html"
    # 如果文件的大小小于20MB，作为附件发送
    if int(record.file_size / 1024 / 1024) <= 20:
        if record:
            msg.attach_file(record.files.path)
    msg.send()
