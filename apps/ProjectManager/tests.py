import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pyminizip
import pymysql

from AuditSQL.settings import MEDIA_ROOT

conn = pymysql.connect(host='10.72.63.128', user='inception_user', password='inception@123com', db='test',
                       charset='utf8')

# 获取标题
try:
    with conn.cursor() as cursor:
        sql = "select uid,mobile,true_name as '姓名',birthday,invite_code,created_at,nickname from xboss_account limit 100"
        cursor.execute(sql)

        c_result = []

        for row in cursor.fetchall():
            c_result.append(row)

    conn.cursorclass = pymysql.cursors.DictCursor
    with conn.cursor() as cursor:
        sql = "select uid,mobile,true_name as '姓名',birthday,invite_code,created_at,nickname from xboss_account limit 100"
        cursor.execute(sql)

        c_title = []

        for key in cursor.fetchone():
            c_title.append(key)

finally:
    conn.close()


wb = Workbook()
wb.encoding = 'utf-8'
ws = wb.active
ws.title = '表格ss'
font = Font(name='Courier', size=14)
align = Alignment(horizontal='right', vertical='center')

ws.append(c_title)
for irow in c_result:
    ws.append(irow)

for row in range(1, ws.max_row + 1):
    for column in range(1, ws.max_column + 1):
        ws.cell(row=row, column=column).font = font
        ws.cell(row=row, column=column).alignment = align
        ws.row_dimensions[row].height = 18
        column_alias = ws.cell(row=row, column=column).column
        ws.column_dimensions[f'{column_alias}'].width = 15

wb.save(f'{MEDIA_ROOT}/files/dadas.xlsx')
pyminizip.compress_multiple([f'{MEDIA_ROOT}/files/dadas.xlsx'], f'{MEDIA_ROOT}/files/dadas.xlsx.zip', "1233", 4)
