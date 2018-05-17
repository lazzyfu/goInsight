# -*- coding:utf-8 -*-
# edit by fuzongfei

import json
import os
import pyminizip
import random
import string
import time

import pymysql
from asgiref.sync import async_to_sync
from celery import shared_task
from celery.result import AsyncResult
from channels.layers import get_channel_layer
from django.core.files import File
from django.core.mail import EmailMessage
from django.db.models import F
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from AuditSQL.settings import EMAIL_FROM
from project_manager.inception.inception_api import IncepSqlCheck
from project_manager.models import AuditContents, OlAuditContentsReply, IncepMakeExecTask, \
    DomainName, InceptionHostConfig, OlDataExportDetail, OlAuditDetail, ExportFiles
from project_manager.utils import update_tasks_status
from user_manager.utils import GetEmailAddr

channel_layer = get_channel_layer()


@shared_task
def send_commit_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(AuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operator')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向_commit_mail.html渲染data数据
    domain_name = ''
    if DomainName.objects.filter().first():
        domain_name = DomainName.objects.get().domain_name
    data = AuditContents.objects.annotate(group_name=F('group__group_name')).get(pk=latest_id)
    if data.audit_type == '0':
        detail = OlAuditDetail.objects.get(ol=latest_id)
    else:
        detail = OlDataExportDetail.objects.get(ol=latest_id)
    email_html_body = render_to_string('_send_commit_mail.html', {
        'data': data,
        'detail': detail,
        'domain_name': domain_name
    })

    # 发送邮件
    msg = EmailMessage(subject=data.title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       )
    msg.content_subtype = "html"

    msg.send()


@shared_task
def send_verify_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(AuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operator')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向mail_template.html渲染data数据
    data = AuditContents.objects.get(pk=latest_id)
    email_html_body = render_to_string('_send_verify_mail.html', {
        'data': data,
        'type': kwargs.get('type'),
        'user_role': kwargs.get('user_role'),
        'username': kwargs.get('username'),
        'addition_info': kwargs.get('addition_info')
    })

    # 发送邮件
    headers = {'Reply: ': receiver}
    title = 'Re: ' + data.title
    msg = EmailMessage(subject=title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       headers=headers)
    msg.content_subtype = "html"
    msg.send()


@shared_task
def send_reply_mail(**kwargs):
    latest_id = kwargs['latest_id']
    reply_id = kwargs['reply_id']
    userinfo = GetEmailAddr(AuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operator')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    title = AuditContents.objects.get(pk=reply_id).title
    reply_record = OlAuditContentsReply.objects.get(pk=latest_id)

    # 向mail_template.html渲染data数据
    email_html_body = render_to_string('_send_reply_mail.html', {
        'reply_record': reply_record,
        'username': kwargs['username'],
    })

    # 发送邮件
    headers = {'Reply: ': receiver}
    title = 'Re: ' + title
    msg = EmailMessage(subject=title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       headers=headers)
    msg.content_subtype = "html"
    msg.send()


"""
status = 0: 推送执行结果
status = 1: 推送执行进度
status = 2: 推送inception processlist
"""


@shared_task
def get_osc_percent(task_id):
    """实时获取pt-online-schema-change执行进度"""
    task = AsyncResult(task_id)

    while task.state in ('PENDING', 'STARTED', 'PROGRESS'):
        while task.state == 'PROGRESS':
            user = task.result.get('user')
            host = task.result.get('host')
            database = task.result.get('database')
            sqlsha1 = task.result.get('sqlsha1')

            sql = f"inception get osc_percent '{sqlsha1}'"
            incep_of_audit = IncepSqlCheck(sql, host, database, user)

            # 执行SQL
            incep_of_audit.run_status(1)

            # 每1s获取一次
            time.sleep(1)
        else:
            continue


@shared_task(bind=True)
def incep_async_tasks(self, id=None, user=None, sql=None, sqlsha1=None, host=None, database=None, exec_status=None,
                      backup=None):
    # 更新任务状态为：PROGRESS
    self.update_state(state="PROGRESS", meta={'user': user, 'host': host, 'database': database, 'sqlsha1': sqlsha1})

    incep_of_audit = IncepSqlCheck(sql, host, database, user)

    # 执行SQL
    exec_result = incep_of_audit.run_exec(0, backup)

    # 更新任务进度
    update_tasks_status(id=id, exec_result=exec_result, exec_status=exec_status)


@shared_task
def stop_incep_osc(user, id=None, celery_task_id=None):
    obj = IncepMakeExecTask.objects.get(id=id)
    host = obj.dst_host
    database = obj.dst_database

    exec_status = None
    if obj.exec_status == '2':
        sqlsha1 = obj.sqlsha1
        exec_status = 0
    elif obj.exec_status == '3':
        sqlsha1 = obj.rollback_sqlsha1
        exec_status = 1

    sql = f"inception stop alter '{sqlsha1}'"

    # 执行SQL
    task = AsyncResult(celery_task_id)
    if task.state == 'PROGRESS':
        incep_of_audit = IncepSqlCheck(sql, host, database, user)
        incep_of_audit.run_status(0)

        # 更新任务进度
        update_tasks_status(id=id, exec_status=exec_status)


@shared_task
def make_export_file(user, id):
    obj = AuditContents.objects.get(pk=id)
    data = OlDataExportDetail.objects.get(ol=id)

    config = InceptionHostConfig.objects.get(host=obj.host, is_enable=0)

    conn = pymysql.connect(host=config.host,
                           user=config.user,
                           password=config.password,
                           port=config.port,
                           database=obj.database,
                           charset="utf8")

    # 创建目录和生成文件名
    tmp_file = f'media/tmp/{obj.title}.{data.file_format}'
    zip_file = tmp_file + '.zip'

    try:
        sql = data.contents
        # 获取内容
        with conn.cursor() as cursor:
            cursor.execute(sql)

            c_result = []

            for row in cursor.fetchall():
                c_result.append(row)

        # 获取标题
        conn.cursorclass = pymysql.cursors.DictCursor
        with conn.cursor() as cursor:
            cursor.execute(sql)

            c_title = []

            for key in cursor.fetchone():
                c_title.append(key)

        wb = Workbook()
        wb.encoding = f'{data.file_coding}'
        ws = wb.active
        font = Font(name='Courier', size=14)
        align = Alignment(horizontal='right', vertical='center')

        # 写入数据
        ws.append(c_title)
        for irow in c_result:
            ws.append(irow)

        # 设置表格的样式
        for row in range(1, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                ws.cell(row=row, column=column).font = font
                ws.cell(row=row, column=column).alignment = align
                ws.row_dimensions[row].height = 18
                column_alias = ws.cell(row=row, column=column).column
                ws.column_dimensions[f'{column_alias}'].width = 15

        wb.save(tmp_file)

        # 压缩并加密，随机生成12位长度的字符串
        salt = ''.join(random.sample(string.ascii_letters + string.digits, 12))
        pyminizip.compress_multiple([tmp_file], zip_file, salt, 4)

        # 将文件转换为File对象，便于存储
        with open(zip_file, 'rb') as f:
            myfile = File(f)
            ExportFiles.objects.create(
                export=data,
                file_name=obj.title + '.xlsx.gz',
                file_size=int(myfile.size / 1024),
                files=myfile,
                encryption_key=salt
            )

        latest_id = ExportFiles.objects.latest('id').id

        # 删除文件，保留加密和压缩后的文件
        os.remove(tmp_file)
        os.remove(zip_file)

        # 更新进度
        OlDataExportDetail.objects.filter(ol=id).update(progress='2')

        send_data_export_mail.delay(latest_id=latest_id, title=obj.title)
    except conn.InternalError as err:
        pull_msg = {'status': 2, 'msg': str(err)}
        OlDataExportDetail.objects.filter(ol=id).update(progress='0')
        async_to_sync(channel_layer.group_send)(user, {"type": "user.message",
                                                       'text': json.dumps(pull_msg)})

    finally:
        conn.close()


@shared_task
def send_data_export_mail(**kwargs):
    latest_id = kwargs['latest_id']
    userinfo = GetEmailAddr(AuditContents, latest_id)

    receiver = userinfo.get_user_email('proposer', 'verifier', 'operator')
    cc = userinfo.get_contact_email()
    bcc = userinfo.get_bcc_email()

    # 向mail_template.html渲染data数据
    domain_name = ''
    if DomainName.objects.filter().first():
        domain_name = DomainName.objects.get().domain_name
    detail = ExportFiles.objects.get(pk=latest_id)
    title = kwargs.get('title')
    email_html_body = render_to_string('_send_data_export_mail.html', {
        'data': detail,
        'domain_name': domain_name,
        'file_size_mb': int(detail.file_size / 1024),
    })

    # 发送邮件
    headers = {'Reply: ': receiver}
    title = 'Re: ' + title
    msg = EmailMessage(subject=title,
                       body=email_html_body,
                       from_email=EMAIL_FROM,
                       to=receiver,
                       cc=cc,
                       bcc=bcc,
                       headers=headers)
    msg.content_subtype = "html"
    # 如果文件的大小小于20MB，作为附件发送
    if int(detail.file_size / 1024) <= 20:
        msg.attach_file(detail.files.path)
    msg.send()
